import openpyxl


def rowcount(file, sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.max_row


def readdata(file, sheetname, row, colno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row, colno).value


def writedata(file, sheetname, row, colno, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row, colno).value == data
    book.save(file)
